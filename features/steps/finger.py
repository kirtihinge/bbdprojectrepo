import time

import openpyxl
from openpyxl.styles import PatternFill, Font
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By

from behave import *
from selenium import webdriver

from features.steps.commonexcelstep import getCellData, failStatus, passStatus, writeCellData, getRowCount


@given(u'open chrome browser')
def step_impl(context):
    context.driver = webdriver.Chrome(executable_path="C:\selenium driver\chromedriver_win32\chromedriver.exe")


@when(u'enter url')
def step_impl(context):
    context.driver.get('http://mitintech.com/admin/login/?next=/admin/')
    context.driver.maximize_window()
    time.sleep(5)


# @when(u'enter username "{username}" and password "{password}"')
# def step_impl(context, username, password):
#     context.driver.find_element(By.NAME, 'username').send_keys(username)
#     context.driver.find_element(By.NAME, 'password').send_keys(password)

@when(u'enter username "kirtihinge" and password "prutha20"')
def step_impl(context):
    context.driver.find_element(By.NAME, 'username').send_keys("kirtihinge")
    context.driver.find_element(By.NAME, 'password').send_keys("prutha20")


@then(u'click on login button')
def step_impl(context):
    # context.driver.find_element(By.XPATH, '//input[@type="submit"]').click()
    context.driver.find_element(By.XPATH, '//input[@value="Log in"]').click()


# @then(u'Login page should be displayed')
# def step_impl(context):
#     title = context.driver.title
#     assert title == "Five Fingers Admin | Five Fingers Admin Site"


@given(u'Click on book series')
def step_impl(context):
    time.sleep(3)
    context.driver.find_element(By.PARTIAL_LINK_TEXT, 'Book Series').click()


@when(u'Click on Add Book Series')
def step_impl(context):
    context.driver.find_element(By.XPATH, '//a[@href="/admin/master/tblbookseries/add/"]').click()
    time.sleep(5)


# @when(u'Enter Book series name')
# def step_impl(context):
#     context.driver.find_element(By.NAME, "name").send_keys("perimeter")


# @when(u'Click on Save')
# def step_impl(context):
#     context.driver.find_element(By.NAME, "_save").click()


# @then(u'check book series name should be added')
# def step_impl(context):
#     title = context.driver.title
#     assert title == "Select Book Serie to change | Five Fingers admin site"
#
#
# @when(u'Click on Add Another subject')
# def step_impl(context):
#     context.driver.find_element(By.XPATH, '//img[@src="/static/admin/img/icon-changelink.svg"]').click()
#
#
# @when(u'Click on Change another subject')
# def step_impl(context):
#     # context.driver.find_element(By.XPATH,'//img[@src="/static/admin/img/icon-changelink.svg"]').click()
#     context.child = context.driver.window_handles[1]
#     context.driver.switch_to.window(context.child)
#
#
# @when(u'clear data')
# def step_impl(context):
#     context.driver.find_element(By.NAME, "name").clear()
#     context.driver.find_element(By.NAME, "description").clear()
#
#
# @when(u'Enter name and description')
# def step_impl(context):
#     time.sleep(5)
#     context.driver.find_element(By.NAME, "name").send_keys("sci")
#     context.driver.find_element(By.NAME, "description").send_keys("this is science")
#     context.driver.find_element(By.NAME, "_save")
#
#
# @then(u'Add book series page should be displayed')
# def step_impl(context):
#     title = context.driver.title
#     assert title == "Add Subject serie|Five Finger admin site"
#
#
# @when(u'select subject and enter name')
# def step_impl(context):
#     context.driver.find_element(By.NAME, 'subject').send_keys('maths')
#     context.driver.find_element(By.NAME, 'name').send_keys('speed time distance')
#
#
# @when(u'click on Save and continue editing')
# def step_impl(context):
#     time.sleep(5)
#     context.driver.find_element(By.NAME, '_continue').click()
#
#
# @when(u'click on delete button')
# def step_impl(context):
#     time.sleep(5)
#     context.driver.find_element(By.CLASS_NAME, 'deletelink').click()
#
#
# @when(u'click on yes I\'m sure button')
# def step_impl(context):
#     time.sleep(5)
#     context.driver.find_element(By.XPATH, '//input[@value="Yes, I’m sure"]').click()
#     time.sleep(5)
#
#
# @then(u'book series name should be deleted')
# def step_impl(context):
#     title = context.driver.title
#     assert title == "Select Book Serie to change|Five Finger admin site"
#
#
# @when(u'click on No,take me back button')
# def step_impl(context):
#     context.driver.find_element(By.XPATH, '//a[@class="button cancel-link"]').click()
#     time.sleep(3)
#
#
# @when(u'click on Add')
# def step_impl(context):
#     context.driver.find_element(By.XPATH, '//img[@src="/static/admin/img/icon-changelink.svg"]').click()
#
#
# @when(u'select book series')
# def step_impl(context):
#     context.driver.find_element(By.XPATH, '//input[@value="94"]').click()
#     time.sleep(3)
#
#
# @when(u'select Action in dropdown')
# def step_impl(context):
#     context.driver.find_element(By.NAME, 'action').click()
#     context.driver.find_element(By.XPATH, '//select[@name="action"]/option[2]').click()
#
#
# @when(u'click on Go button')
# def step_impl(context):
#     context.driver.find_element(By.NAME, 'index').click()
#
#
# @when(u'select more than book series')
# def step_impl(context):
#     context.driver.find_element(By.XPATH, '//input[@value="93"]').click()
#     context.driver.find_element(By.XPATH, '//input[@value="94"]').click()
#
#
# @when(u'click on History')
# def step_impl(context):
#     context.driver.find_element(By.CLASS_NAME, 'historylink').click()
#
#
# @then(u'history of book series')
# def step_impl(context):
#     context.driver.find_element(By.CLASS_NAME, 'historylink').click()

# def xyz(sno):
#     Subject = ''
#     Name = ''
#     time.sleep(3)
#     path = "C:\KIRTI\MANUAL testing\write da.xlsx"
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook["Book"]
#     rows = sheet.max_row
#     for i in range(1, rows):
#
#         if sno == sheet.cell(row=i + 1, column=1).value:
#             Subject = sheet.cell(row=i + 1, column=2).value
#             Name = sheet.cell(row=i + 1, column=3).value
#         # i = i + 1
#     return Subject, Name
#------------------------------------------------------------------------
def xyz(sno):
    Subject = ''
    Name = ''
    time.sleep(3)
    path = "C:\KIRTI\MANUAL testing\write da.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Book"]
    rows = getRowCount("C:\KIRTI\MANUAL testing\write da.xlsx", "Book")
    for i in range(1, rows):
        a=getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Book', i + 1, 1)
        if sno == a:
            Subject = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Book', i + 1, 2)
            Name = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Book', i + 1, 3)
        # i = i + 1
    return Subject, Name

@when(u'Enter Book series name "{sno}"')
def step_impl(context, sno):
    s, b = xyz(sno)
    time.sleep(5)
    context.driver.find_element(By.NAME, 'subject').send_keys(s)
    context.driver.find_element(By.NAME, 'name').send_keys(b)


@when(u'Click on Save')
def step_impl(context):
    context.driver.find_element(By.NAME, "_save").click()
    time.sleep(3)


@then(u'check book series name should be added "{status}""{sno}"')
def step_impl(context, status, sno):
    title = context.driver.title
    print(title)
    if title.strip() == "Select Book Serie to change | Five Fingers admin site".strip():

        time.sleep(10)
        path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        workbook = openpyxl.load_workbook(path)

        sheet = workbook['Book']
        # cl = sheet['D2']
        # cl.fill = PatternFill("solid", start_color="228B22")
        print(sheet.max_row)

        for i in range(2, sheet.max_row + 1):
            cl = sheet['D' + str(i)]
            s, b = xyz(sno)

            # if sheet.cell(row=i, column=2).value == 'Maths' and sheet.cell(row=i, column=3).value == 'perimeter':
            # if sheet.cell(row=i, column=2).value == s and sheet.cell(row=i, column=3).value == b:
            #     cl.fill = PatternFill("solid", start_color="228B22")
            #     cl.font = Font(color='ffffff')
            #     cl.font = Font(size=10)
            #     cl.font = cl.font.copy(bold=True)
            #     sheet.cell(row=i, column=4).value = "pass"
            #     workbook.save(path)
            if getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Book', i, 2) == s and getCellData(
                    "C:\KIRTI\MANUAL testing\write da.xlsx", 'Book', i, 3) == b:
                status = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Book', i, 4)
                passStatus(cl, status)
                # cl.fill = PatternFill("solid", start_color="228B22")
                # cl.font = Font(color='ffffff')
                # cl.font = Font(size=10)
                # cl.font = cl.font.copy(bold=True)
                # sheet.cell(row=i, column=4).value = "pass"
                workbook.save(path)
    # else:
    #     time.sleep(10)
    #     path = "C:\KIRTI\MANUAL testing\write da.xlsx"
    #     workbook = openpyxl.load_workbook(path)
    #     sheet = workbook['Book']
    #     # cl = sheet['D3']
    #     for i in range(2, sheet.max_row):
    #         if sheet.cell(row=i, column=2).value != 'maths' and sheet.cell(row=i, column=3).value != 'perimeter':
    #             sheet.cell(row=i, column=4).value = status
    #             cl = status
    #             cl.value = "Fail"
    #             cl.fill = PatternFill("solid", start_color="A52A2A")
    #             cl.font = Font(color='ffffff')
    #             cl.font = cl.font.copy(bold=True)
    #             cl.font = Font(size=14)
    #             workbook.save(path)
    #
    #             get_title = context.driver.title
    #             # print(get_title)
    #             path = "C:\KIRTI\MANUAL testing\write da.xlsx"
    #             workbook = openpyxl.load_workbook(path)
    #             sheet = workbook['Book']
    #             for i in range(2, sheet.max_row):
    #                 if sheet.cell(row=i, column=2).value != s and sheet.cell(row=i, column=3).value != b:
    #                     sheet.cell(row=i, column=5).value = get_title
    #             # sheet['E3'] = get_title
    #     workbook.save(path)
# -------------------------------------------------------------------------------
    else:
        time.sleep(10)
        path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook['Book']
        # cl = sheet['D3']

        for i in range(2, sheet.max_row):
            cl = sheet['D' + str(i)]
            s, b = xyz(sno)

            if getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Book', i, 2) == s and getCellData(
                        "C:\KIRTI\MANUAL testing\write da.xlsx", 'Book', i, 3) == b:
                    status = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Book', i, 4)
                    passStatus(cl, status)
                # cl.value = "Fail"
                # cl.fill = PatternFill("solid", start_color="A52A2A")
                # cl.font = Font(color='ffffff')
                # cl.font = cl.font.copy(bold=True)
                # cl.font = Font(size=14)
                    workbook.save(path)

                    get_title = context.driver.title
                    path = "C:\KIRTI\MANUAL testing\write da.xlsx"
                    workbook = openpyxl.load_workbook(path)
                    sheet = workbook['Book']
                    writeCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Book', i, 5, get_title)

