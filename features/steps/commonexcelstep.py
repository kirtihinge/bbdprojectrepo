from multiprocessing import context

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.builtins import title

from demo import abc


def getRowCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column


def getCellData(path, sheetName, rNum, cNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rNum, column=cNum).value


def passStatus(cl, status):
    cl.value = status
    cl.value = "pass"
    cl.fill = PatternFill("solid", start_color="228B22")
    cl.font = Font(color='ffffff')
    cl.font = Font(size=10)
    cl.font = cl.font.copy(bold=True)


def failStatus(cl, status):
    cl.fill = PatternFill("solid", start_color="A52A2A")
    cl.font = Font(color='ffffff')
    cl.font = cl.font.copy(bold=True)

    cl.font = Font(size=16)
    cl.value = status
    cl.value = "Fail"


def writeCellData(path, sheetName, rNum, cNum, get_title):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rNum, column=cNum).value = get_title
    workbook.save(path)
