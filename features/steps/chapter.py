import openpyxl
import time

from openpyxl.styles import PatternFill, Font
# from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
#
from behave import *
from selenium import webdriver

from features.steps.commonexcelstep import getCellData, passStatus, failStatus, writeCellData, getRowCount


# @given(u'enter url')
# def step_impl(context):
#     context.driver.get('http://mitintech.com/admin/login/?next=/admin/')


# @given(u'enter username "kirtihinge" and password "prutha20"')
# def step_impl(context):
#     context.driver.find_element(By.NAME, 'username').send_keys('kirtihinge')
#     time.sleep(3)
#     context.driver.find_element(By.NAME, 'password').send_keys('prutha20')
#     time.sleep(3)


# @given(u'click on login button')
# def step_impl(context):
#     context.driver.find_element(By.XPATH, '//input[@type="submit"]').click()
#     context.driver.implicitly_wait(5)
@when(u'click on login button')
def step_impl(context):
    time.sleep(3)
    context.driver.find_element(By.XPATH, '//input[@type="submit"]').click()


@when(u'click on chapter')
def step_impl(context):
    time.sleep(3)
    context.driver.find_element(By.PARTIAL_LINK_TEXT, 'Chapter').click()


#
#
# @when(u'click on add button')
# def step_impl(context):
#     context.driver.find_element(By.XPATH,'//tr[@class="model-tblchapter"]/td[1]/a').click()
#
# @then(u'Add chapter page should page should be opened')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then Add chapter page should page should be opened')
#
#
# @when(u'Enter chapter')
# def step_impl(context):
#     context.driver.find_element(By.NAME, 'name').send_keys('perimeter')
#     context.driver.find_element(By.NAME,'subject').send_keys('maths')
#     context.driver.find_element(By.NAME,'bookseries').send_keys('algebra')
#     context.driver.find_element(By.NAME,'classid').send_keys('v')
#     context.driver.find_element(By.NAME, '_save').click()
# @when(u'click on save button')
# def step_impl(context):
#     context.driver.find_element(By.NAME,'_save').click()

#
# @then(u'chapter should be added sucessfully')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then chapter should be added sucessfully')
#
#
# @when(u'click on save and another button')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on save and another button')
#
#
# @then(u'chapter should be added and add chapter page opened again sucessfully')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then chapter should be added and add chapter page opened again sucessfully')
#
#
# @when(u'click on save and continue editing button')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on save and continue editing button')
#
#
# @then(u'chapter should be added and edit chapter page opened sucessfully')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then chapter should be added and edit chapter page opened sucessfully')
#
#
# @then(u'click on delete option')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then click on delete option')
#
#
# @then(u'click on Yes I\'m sure button')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then click on Yes I\'m sure button')
#
#
# @then(u'click on No take back button')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then click on No take back button')
#
#
# @when(u'select chapter')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When select chapter')
#
#
# @when(u'select Action')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When select Action')
#
#
# @when(u'click on go')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on go')
#
#
# @when(u'click on Yes,I\'m sure')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on Yes,I\'m sure')
#
#
# @then(u'chapter should be deleted')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then chapter should be deleted')
#
#
# @when(u'select action')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When select action')
#
#
# @when(u'click on Yes,I\'m sure')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on Yes,I\'m sure')
@when(u'click on add')
def step_impl(context):
    time.sleep(3)
    # context.driver.find_element(By.XPATH, '//tr[@class="model-tblchapter"]/td[1]/a').click()
    context.driver.find_element(By.XPATH, '//*[@id="nav-sidebar"]/div[2]/table/tbody/tr[2]/td/a').click()


# def ghf(sno):
#     Name = ""
#     Subject = ""
#     Book = ""
#     Classid = ""
#     path = "C:\KIRTI\MANUAL testing\write da.xlsx"
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook["Chapter"]
#     rows = sheet.max_row
#     for i in range(1, rows):
#         if sno == sheet.cell(row=i + 1, column=1).value:
#             Name = sheet.cell(row=i + 1, column=2).value
#             Subject = sheet.cell(row=i + 1, column=3).value
#             Book = sheet.cell(row=i + 1, column=4).value
#             Classid = sheet.cell(row=i + 1, column=5).value
#
#         i = i + 1
#     return Name, Subject, Book, Classid
# ---------------------------------------------------------------------
def ghf(sno):
    Name = ""
    Subject = ""
    Book = ""
    Classid = ""
    path = "C:\KIRTI\MANUAL testing\write da.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Chapter"]
    rows = getRowCount("C:\KIRTI\MANUAL testing\write da.xlsx", "Chapter")
    for i in range(1, rows):
        a = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i + 1, 1)

        if sno == a:
            Name = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i + 1, 2)
            Subject = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i + 1, 3)
            Book = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i + 1, 4)
            Classid = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i + 1, 5)

        i = i + 1
    return Name, Subject, Book, Classid


# ---------------------------------------------------------------------

@when(u'Enter chapter "{sno}"')
def step_impl(context, sno):
    n, s, b, c = ghf(sno)
    context.driver.find_element(By.NAME, 'name').send_keys(n)
    context.driver.find_element(By.NAME, 'subject').send_keys(s)
    context.driver.find_element(By.NAME, 'bookseries').send_keys(b)
    context.driver.find_element(By.NAME, 'classid').send_keys(c)
    context.driver.find_element(By.NAME, '_save').click()


# @then(u'chapter should be added sucessfully "{status}""{sno}"')
# def step_impl(context, status, sno):
#     n, s, b, c = ghf(sno)
#     title = context.driver.title
#     if title == "Select Chapter to change | Five Fingers admin site":
#
#         time.sleep(10)
#         path = "C:\KIRTI\MANUAL testing\write da.xlsx"
#         workbook = openpyxl.load_workbook(path)
#
#         sheet = workbook['Chapter']
#         # cl = sheet['F2']
#         for i in range(2, sheet.max_row):
#             if sheet.cell(row=i, column=2).value == n and sheet.cell(row=i, column=3).value == s:
#                 sheet.cell(row=i, column=4).value = status
#                 cl = sheet
#                 cl.fill = PatternFill("solid", start_color="228B22")
#                 cl.font = Font(color='ffffff')
#                 cl.font = Font(size=10)
#                 cl.font = cl.font.copy(bold=True)
#                 cl.value = "Pass"
#                 time.sleep(3)
#         workbook.save(path)
#
#     else:
#         time.sleep(10)
#         path = "C:\KIRTI\MANUAL testing\write da.xlsx"
#         workbook = openpyxl.load_workbook(path)
#         sheet = workbook['Chapter']
#         cl = sheet['F3']
#         cl.value = "Fail"
#         cl.fill = PatternFill("solid", start_color="A52A2A")
#         cl.font = Font(color='ffffff')
#         cl.font = cl.font.copy(bold=True)
#         cl.font = Font(size=14)
#
#         get_title = context.driver.title
#         # print(get_title)
#         path = "C:\KIRTI\MANUAL testing\write da.xlsx"
#         workbook = openpyxl.load_workbook(path)
#         sheet = workbook['Chapter']
#         sheet['G3'] = get_title
#         workbook.save(path)

# -------------------------------------------------------------------
@then(u'chapter should be added sucessfully "{status}""{sno}"')
def step_impl(context, status, sno):

    title = context.driver.title
    if title == "Select Chapter to change | Five Fingers admin site":

        time.sleep(10)
        path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        workbook = openpyxl.load_workbook(path)

        sheet = workbook['Chapter']
        for i in range(2, sheet.max_row + 1):
            cl = sheet['F' + str(i)]
            n, s, b, c = ghf(sno)
            if getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i, 2) == n and getCellData(
                    "C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i, 3) == s:
                status = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i, 4)
                passStatus(cl, status)
                # cl.fill = PatternFill("solid", start_color="228B22")
                # cl.font = Font(color='ffffff')
                # cl.font = Font(size=10)
                # cl.font = cl.font.copy(bold=True)
                # cl.value = "Pass"
                # time.sleep(3)
                workbook.save(path)

    else:
        time.sleep(10)
        path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook['Chapter']
        for i in range(2, sheet.max_row + 1):
            cl = sheet['F' + str(i)]
            n, s, b, c = ghf(sno)
            if getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i, 2) == n and getCellData(
                    "C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i, 3) == s:
                status = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i, 6)
                failStatus(cl, status)
                get_title = context.driver.title
                # print(get_title)
                path = "C:\KIRTI\MANUAL testing\write da.xlsx"
                workbook = openpyxl.load_workbook(path)
                sheet = workbook['Chapter']
                writeCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Chapter', i, 7, get_title)

        # cl = sheet['F3']
        # cl.value = "Fail"
        # cl.fill = PatternFill("solid", start_color="A52A2A")
        # cl.font = Font(color='ffffff')
        # cl.font = cl.font.copy(bold=True)
        # cl.font = Font(size=14)
        #
        # get_title = context.driver.title
        # # print(get_title)
        # path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        # workbook = openpyxl.load_workbook(path)
        # sheet = workbook['Chapter']
        # sheet['G3'] = get_title
        workbook.save(path)
