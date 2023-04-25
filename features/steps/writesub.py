import time

from selenium.webdriver.common.by import By
import openpyxl
from openpyxl.styles import PatternFill, Font
from behave import *
from selenium import webdriver

from features.steps.commonexcelstep import getRowCount, getColumnCount, getCellData, passStatus, failStatus, \
    writeCellData


# @given(u'enter url')
# def step_impl(context):
#     context.driver.get('http://mitintech.com/admin/login/?next=/admin/')


@given(u'enter username "kirtihinge" and password "prutha20"')
def step_impl(context):
    # time.sleep(3)
    context.driver.find_element(By.NAME, 'username').send_keys('kirtihinge')
    # time.sleep(3)
    context.driver.find_element(By.NAME, 'password').send_keys('prutha20')
    time.sleep(3)


@given(u'click on login button')
def step_impl(context):
    context.driver.find_element(By.XPATH, '//input[@type="submit"]').click()
    context.driver.implicitly_wait(5)


@when(u'click on add subject')
def step_impl(context):
    # context.driver.find_element(By.XPATH, '//tr[@class="model-tblsubject"]/td[1]/a').click()
    context.driver.find_element(By.XPATH, '//*[@id="content-main"]/div[2]/table/tbody/tr[4]/td[1]/a').click()
    # time.sleep(3)


# def abc(sno):
#     Name = ''
#     Description = ''
#
#     path = "C:\KIRTI\MANUAL testing\write da.xlsx"
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook["Subject"]
#     rows = sheet.max_row
#     for i in range(1, rows):
#         if sno == sheet.cell(row=i + 1, column=1).value:
#             Name = sheet.cell(row=i + 1, column=2).value
#             Description = sheet.cell(row=i + 1, column=3).value
#             print(Name, ' --->', Description)
#         i = i + 1
#     return Name, Description
# --------------------------------------------------------------------
def abc(sno):
    Name = ''
    Description = ''

    path = "C:\KIRTI\MANUAL testing\write da.xlsx"
    rows = getRowCount("C:\KIRTI\MANUAL testing\write da.xlsx", "Subject")
    column = getColumnCount("C:\KIRTI\MANUAL testing\write da.xlsx", "Subject")

    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Subject"]
    rows=getRowCount("C:\KIRTI\MANUAL testing\write da.xlsx", "Subject")
    # print(rows)
    # print(column)
    for i in range(1, rows):
        a = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i + 1, 1)
        if sno == a:
            Name = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i + 1, 2)
            Description = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i + 1, 3)
            print(Name, ' --->', Description)
        i = i + 1
    return Name, Description


# --------------------------------------------------------------------

# @when(u'enter "{sno}" with Name and Description')
# def step_impl(context, sno):
#     context.driver.find_element(By.NAME, 'name').send_keys(subjectpom.Name)
#     context.driver.find_element(By.NAME, 'description').send_keys(subjectpom.Description)

@when(u'enter "{sno}" with Name and Description')
def step_impl(context, sno):
    n, d = abc(sno)
    print(n, d)
    # time.sleep(10)
    context.driver.find_element(By.NAME, 'name').send_keys(n)
    context.driver.find_element(By.NAME, 'description').send_keys(d)


# @when(u'enter "{sno}" with Name and Description')
# def step_impl(context, sno):
#     Name = ''
#     Description = ''
#
#     path = "C:\KIRTI\MANUAL testing\write da.xlsx"
#     workbook = openpyxl.load_workbook(path)
#     # sheet = workbook["Subject"]
#     sheet = workbook["Sheet4"]
#     rows = sheet.max_row
#     for i in range(1, rows):
#         if sheet.cell(row=i+1, column=1).value == sno:
#             Name = sheet.cell(row=i + 1, column=2).value
#             Description = sheet.cell(row=i + 1, column=3).value
#         i = i + 1
#     context.driver.find_element(By.NAME, 'name').send_keys(Name)
#     context.driver.find_element(By.NAME, 'description').send_keys(Description)


@then(u'click on Save button state')
def step_impl(context):
    context.driver.find_element(By.NAME, '_save').click()
    # time.sleep(3)


@then(u'subject page should be displayed "{status}""{sno}"')
def step_impl(context, status, sno):
    title = context.driver.title
    if title.strip() == "Select Subject to change | Five Fingers admin site".strip():
        # print(title)
        time.sleep(3)
        path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook['Subject']
        for i in range(2, sheet.max_row + 1):
            cl = sheet['D' + str(i)]
            n, d = abc(sno)
            if getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i, 2) == n and getCellData(
                    "C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i, 3) == d:
                status = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i, 4)
                passStatus(cl, status)
            #         cl.value = 'pass'
            #         cl.fill = PatternFill("solid", start_color="228B22")
            #         cl.font = Font(color='ffffff')
            #         cl.font = Font(size=10)
            #         cl.font = cl.font.copy(bold=True)
                workbook.save(path)
    else:
        time.sleep(10)
        path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook['Subject']
        for i in range(2, sheet.max_row + 1):
            cl = sheet['D' + str(i)]
            print(sheet.cell(row=i, column=2).value)
            print(sheet.cell(row=i, column=3).value)
            n, d = abc(sno)
            print(n, '---', d)
            if getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i, 2) == n and getCellData(
                    "C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i, 3) == d:
                status = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i, 4)
                failStatus(cl, status)
                #         cl.fill = PatternFill("solid", start_color="A52A2A")
                #         cl.font = Font(color='ffffff')
                #         cl.font = cl.font.copy(bold=True)
                #
                #         cl.font = Font(size=16)
                #         sheet.cell(row=i, column=4).value = "Fail"
                workbook.save(path)
            # get_title = context.driver.title
            # path = "C:\KIRTI\MANUAL testing\write da.xlsx"
            # workbook = openpyxl.load_workbook(path)
            # # sheet = workbook['Subject']
            # sheet.cell(row=i, column=5).value = get_title
            # t = get_title
            # workbook.save(path)
            get_title = context.driver.title
            path = "C:\KIRTI\MANUAL testing\write da.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook['Subject']

            # getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', )
            writeCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i, 5, get_title)
