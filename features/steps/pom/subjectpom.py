import openpyxl

Name = ''
Description = ''
path = "C:\KIRTI\MANUAL testing\write da.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet4"]
rows = sheet.max_row
for i in range(1, rows):
    if i == sheet.cell(row=i, column=1).value:
        Name = sheet.cell(row=i + 1, column=2).value
        Description = sheet.cell(row=i + 1, column=3).value
