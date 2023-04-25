import time
# from datetime import time

import openpyxl
from openpyxl.styles import PatternFill, Font
# from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
#
from behave import *
from selenium import webdriver

from features.steps.commonexcelstep import getCellData, passStatus, failStatus, writeCellData, getRowCount


# @given(u'enter url')
# def step_impl(context):
#     context.driver.get('http://mitintech.com/admin/login/?next=/admin/')
#
#
# @given(u'enter username "kirtihinge" and password "prutha20"')
# def step_impl(context):
#     context.driver.find_element(By.NAME, 'username').send_keys('kirtihinge')
#     time.sleep(3)
#     context.driver.find_element(By.NAME, 'password').send_keys('prutha20')
#     time.sleep(3)
#
#
# @given(u'click on login button')
# def step_impl(context):
#     context.driver.find_element(By.XPATH, '//input[@type="submit"]').click()
#     context.driver.implicitly_wait(5)


# @when(u'click on class')
# def step_impl(context):
#     context.driver.find_element(By.PARTIAL_LINK_TEXT, 'Class').click()


@when(u'click on class')
def step_impl(context):
    context.driver.find_element(By.PARTIAL_LINK_TEXT, 'Class').click()
    time.sleep(5)


@when(u'click on add button')
def step_impl(context):
    context.driver.find_element(By.XPATH, '//*[@id="nav-sidebar"]/div[2]/table/tbody/tr[3]/td/a').click()
    time.sleep(3)


# @then(u'Add class page should page should be opened')
# def step_impl(context):


# @when(u'Enter class')
# def step_impl(context):


# @when(u'click on save button')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on save button')
#
#
# @then(u'class should be added sucessfully')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then class should be added sucessfully')
#
#
# @when(u'click on save and another button')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on save and another button')
#
#
# @then(u'class should be added and add class page opened again sucessfully')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then class should be added and add class page opened again sucessfully')
#
#
# @when(u'click on save and continue editing button')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on save and continue editing button')
#
#
# @then(u'class should be added and edit class page opened sucessfully')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then class should be added and edit class page opened sucessfully')
#
#
# @then(u'click on delete option')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then click on delete option')
#
#
# @then(u'click on Yes I\'m sure button')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then click on Yes I\'m sure button')
#
#
# @then(u'click on No take back button')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then click on No take back button')
#
#
# @when(u'select class')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When select class')
#
#
# @when(u'select Action')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When select Action')
#
#
# @when(u'click on go')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on go')
#
#
# @when(u'click on Yes,I\'m sure')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on Yes,I\'m sure')
#
#
# @then(u'class should be deleted')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: Then class should be deleted')
#
#
# @when(u'select classes')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When select classes')
#
#
# @when(u'select action')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When select action')
#
#
# @when(u'click on Yes,I\'m sure')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on Yes,I\'m sure')
#
#
# @then(u'classes should be deleted')
# def step_impl(context):


# @when(u'Add class page should page should be opened')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When Add class page should page should be opened')
#
#
# @when(u'Enter class')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When Enter class')
#
#
# @when(u'click on save button')
# def step_impl(context):
#     raise NotImplementedError(u'STEP: When click on save button')
#
#
# @then(u'class should be added sucessfully')
# def step_impl(context):
# def cdf(sno):
#     Name = ""
#
#     path = "C:\KIRTI\MANUAL testing\write da.xlsx"
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook["Class"]
#     rows = sheet.max_row
#     for i in range(1, rows):
#         if sno == sheet.cell(row=i + 1, column=1).value:
#             Name = sheet.cell(row=i + 1, column=2).value
#         i = i + 1
#     return Name
# ------------------------------------------------------------------
def cdf(sno):
    Name = ""

    path = "C:\KIRTI\MANUAL testing\write da.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Class"]
    rows = getRowCount("C:\KIRTI\MANUAL testing\write da.xlsx", "Class")
    for i in range(1, rows):
        a = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Class', i + 1, 1)
        if sno == a:
            Name = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Class', i + 1, 2)
        i = i + 1
    return Name


# ------------------------------------------------------------------------
@when(u'Enter class with "{sno}"')
def step_impl(context, sno):
    n = cdf(sno)
    context.driver.find_element(By.NAME, 'name').send_keys(n)


@when(u'click on save button')
def step_impl(context):
    context.driver.find_element(By.NAME, '_save').click()


@then(u'class should be added with "{status}"')
def step_impl(context, status, sno):
    title = context.driver.title
    if title == "Select Class to change | Five Fingers admin site":

        time.sleep(10)
        path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        workbook = openpyxl.load_workbook(path)

        sheet = workbook['Class']
        for i in range(2, sheet.max_row + 1):
            cl = sheet['C' + str(i)]
            n = cdf(sno)
            if getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i, 2) == n:
                status = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i, 3)
                passStatus(cl, status)
        # cl = sheet['C2']
        # cl.fill = PatternFill("solid", start_color="228B22")
        # cl.font = Font(color='ffffff')
        # cl.font = Font(size=10)
        # cl.font = cl.font.copy(bold=True)
        # cl.value = "Pass"
                workbook.save(path)

    else:
        time.sleep(10)
        path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook['Class']
        for i in range(2, sheet.max_row + 1):
            cl = sheet['C' + str(i)]
            n = cdf(sno)
            if getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i, 2) == n:
                status = getCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Subject', i, 3)
                failStatus(cl, status)
        # cl = sheet['C3']
        # cl.value = "Fail"
        # cl.fill = PatternFill("solid", start_color="A52A2A")
        # cl.font = Font(color='ffffff')
        # cl.font = cl.font.copy(bold=True)
        # cl.font = Font(size=14)

        # get_title = context.driver.title
        # # print(get_title)
        # path = "C:\KIRTI\MANUAL testing\write da.xlsx"
        # workbook = openpyxl.load_workbook(path)
        # sheet = workbook['Class']
        # sheet['E3'] = get_title
        # workbook.save(path)
            get_title = context.driver.title
            # print(get_title)
            path = "C:\KIRTI\MANUAL testing\write da.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook['Class']
            writeCellData("C:\KIRTI\MANUAL testing\write da.xlsx", 'Class', i, 4, get_title)
