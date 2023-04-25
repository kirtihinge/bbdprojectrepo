import openpyxl
def abc(sno):
    Name = ''
    Description = ''

    path = "C:\KIRTI\MANUAL testing\write da.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Subject"]
    rows = sheet.max_row
    for i in range(1, rows):
        if sno == sheet.cell(row=i+1, column=1).value:
            Name = sheet.cell(row=i + 1, column=2).value
            Description = sheet.cell(row=i + 1, column=3).value
            print(Name, ' --->', Description)
        i = i + 1
    return Name, Description


print(abc("A1"))